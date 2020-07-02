from playsound import playsound
import time as tm
import pyttsx3
import datetime
import openpyxl
import threading
import os
import random
import logging

"""
1. write a method which will open tkinter window for taking task input once a day in morning.
2. Using sqllite store everyday data of expense and tasks.
"""

logging.basicConfig(filename='audio.log', level=logging.DEBUG, filemode='w', format='%(asctime)s:%(levelname)s:%(message)s')
threadlock = threading.Lock()
event = threading.Event()


def speak(bolo):
	event.set()
	threadlock.acquire()
	engine = pyttsx3.init()
	engine.say(bolo)
	engine.runAndWait()
	engine.stop()
	threadlock.release()
	event.clear()

def task_reminder():
	logging.info("Task reminder started.")
	wb = openpyxl.load_workbook("task_reminder.xlsx")
	sheets = wb.sheetnames
	sheet = wb[sheets[len(sheets)-1]]
	total_rows = sheet.max_row
	logging.debug("Sheet {} has {} rows.".format(sheets[len(sheets)-1], total_rows))
	while True:
		for i in range(total_rows):
			task = sheet.cell(row=i+1, column=1).value
			time = int(sheet.cell(row=i+1, column=2).value)
			hr = "AM"
			if time > 12:
				time = time-12
				hr = hr.replace("AM","PM")
			elif time == 12:
				hr = hr.replace("AM", "PM")
			current_time = datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p").split()
			current_hour = current_time[4].split(":")[0]
			current_min = int(current_time[4].split(":")[1][:2])
			current_clock = current_time[4].split(":")[1][-2:]
			if time == int(current_hour) and current_clock == hr:
				while(event.is_set()):
					tm.sleep(5)
				if current_min in [10,20,30]:
					logging.info("Have you started working on {} yet as it should have started at {} {}".format(task, time, hr))
					speak("Have you started working on {} yet as it should have started at {} {}".format(task, time, hr))
					tm.sleep(59)
				"""
				count = 0
				while(count < 3):
						while(event.is_set()):
							tm.sleep(5)
						print("Have you started working on {} yet as it should have started at {} {}".format(task, time, hr))
						speak("Have you started working on {} yet as it should have started at {} {}".format(task, time, hr))
						tm.sleep(600)
						count+=1
					"""	

def time_reminder():
	logging.info("Time reminder started.")
	files = os.listdir('C:\\Users\\HIMANSHU\\Desktop\\songs')
	i=0
	total_time = 1500
	david = 'HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-US_DAVID_11.0'
	hazel = 'HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-GB_HAZEL_11.0'
	zira = 'HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-US_ZIRA_11.0'
	while True:
		if i != 0:
			speak("25 minutes are over, drink water, take a short walk and perform eye exercise")		
			#playsound('C:\\Users\\HIMANSHU\\Desktop\\audio\\audio.mp3')
			tm.sleep(3)
			song = random.choice(files)
			print("Playing: "+song)
			event.set()
			playsound('C:\\Users\\HIMANSHU\\Desktop\\songs\\'+song)
			event.clear()
		sleeping(total_time)
		i = i+1
	
	
def sleeping(total_time):
	logging.debug("Inside sleeping")
	time_sleep = 0
	while(time_sleep < total_time):
		tm.sleep(300)
		if time_sleep in [0, 600, 1200]:
			cur_time = datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p").split()
			#speak("5 minutes are over, time is "+cur_time[4])
			speak("Train your mind, 25 minutes study, 10 minutes break, no youtube")
			#playsound('C:\\Users\\HIMANSHU\\Desktop\\audio\\5mins.mp3')
		else:
			while(event.is_set()):
				tm.sleep(2)
			#playsound('C:\\Users\\HIMANSHU\\Desktop\\audio\\5mins.mp3')
			speak("Train your mind, 25 minutes study, 10 minutes break, no youtube")
		print("Completed 5 mins at: "+datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p"))
		time_sleep = time_sleep+300


def task_entry():
	logging.info("Task entry started.")
	current_time = datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p").split()
	sheet_name = "{}{}.{}".format(current_time[1], current_time[2], current_time[3])
	try:
		wb = openpyxl.load_workbook("task_reminder.xlsx")
		sheets = wb.sheetnames
	except Exception as e:
		logging.error(e)
	
	if sheet_name not in sheets:
		logging.debug("creating sheet: "+sheet_name)
		speak("Please provide today's task details.")
		wb.create_sheet(sheet_name)
		wkbook = wb[sheet_name]
		total_task = {}
		def task_details():
			task = input("Enter task details:")
			time = input("Enter time (24 hr format:)")
			total_task[task] = time
		task_details()
		while "y" in input("Want to enter more details:"):
			task_details()
	
		row = 0
		for task in total_task:
			wkbook.cell(row = row+1, column=1).value = task
			wkbook.cell(row = row+1, column=2).value = total_task[task]
			row+=1
		wb.save("task_reminder.xlsx")
		logging.info("Task details saved.")
		speak("Your task details for today has been noted. Thank you and have a good day.")
	else:
		speak("Great, you have already provided the task details for today.")
		logging.info("Task details already provided.")


def expense_entry():
	logging.info("Expense entry started.")
	while True:
		current_time = datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p").split()
		current_hour = int(current_time[4].split(":")[0])
		current_clock = current_time[4].split(":")[1][-2:]
		
		if current_hour == 11 and current_clock == 'PM':
			sheet_name = "{}{}.{}".format(current_time[1], current_time[2], current_time[3])
			try:
				wb = openpyxl.load_workbook("expenses.xlsx")
				sheets = wb.sheetnames
			except Exception as e:
				logging.error(e)
			if sheet_name not in sheets:
				speak("Please provide today's expense details.")
				logging.debug("creating sheet: "+sheet_name)
				wb.create_sheet(sheet_name)
				wkbook = wb[sheet_name]
				all_expense = {}
				def expense_details():
					item = input("Enter item details (add respective suffix like food):")
					amount = input("Enter amount:")
					all_expense[item] = amount
				expense_details()
				while "y" in input("Want to enter more details:"):
					expense_details()
		
				row = 0
				for item in all_expense:
					wkbook.cell(row = row+1, column=1).value = item
					wkbook.cell(row = row+1, column=2).value = all_expense[item]
					row+=1
				wb.save("expenses.xlsx")
				logging.debug("Expense details saved.")
				speak("Your expense details for today has been noted. Thank you and have a good day.")
				exit()
			else:
				speak("Great, you have already provided the expense details for today.")
				logging.debug("Expense details already provided.")
				exit()
		else:
			tm.sleep(1200)
			logging.info("sleeping for 5 minutes as time is {}{}".format(current_hour, current_clock))


def daily_routine_entry():
	logging.info("Daily routine started.")
	while True:
		current_time = datetime.datetime.now().strftime("%A, %d. %B %Y %I:%M%p").split()
		sheet_name = "{}{}.{}".format(current_time[1], current_time[2], current_time[3])
		logging.info("Sheet name: "+sheet_name)
		tm.sleep(1200)
		speak("Please, provide details of your last 20 minutes ")
		try:
			wb = openpyxl.load_workbook('daily_routine.xlsx')
		except Exception as e:
			logging.error(e)
		sheets = wb.sheetnames
		if sheet_name not in sheets:
			logging.debug(sheet_name+" not in sheet so creating.")
			wb.create_sheet(sheet_name)	
		logging.debug(sheet_name+" found in daily_routine.xlsx so appending.")
		workbook = wb[sheet_name]
		total_rows = workbook.max_row
		logging.debug("Total rows in sheet "+ str(total_rows))
		num_entry = int(input("How many tasks you want to add(if you are continuing your previous task then type 0):"))
		logging.debug("Total entry to be added: "+str(num_entry))
		for i in range(1,num_entry+1):
			task = input("Enter the task name:")
			start_time = input("Enter start time in 24hr format:")
			end_time = input("Enter end time in 24hr format:")
			workbook.cell(row = total_rows+i, column=1).value = task
			workbook.cell(row = total_rows+i, column=2).value = start_time
			workbook.cell(row = total_rows+i, column=3).value =	end_time
			logging.debug("Adding {} ,{} ,{} into sheet.".format(task, start_time, end_time))
		try:
			wb.save('daily_routine.xlsx')
			logging.debug("Data saved into sheets.")
		except Exception as e:
			logging.error(e)			
def main():
	task_entry()
	
	t1 = threading.Thread(target=task_reminder) 
	t2 = threading.Thread(target=time_reminder) 
	t3 = threading.Thread(target=expense_entry)
	t4 = threading.Thread(target=daily_routine_entry)
	
	speak("Fight for it, or let it go!")
 
	t1.start()
	logging.debug("Thread t1 started.")
	t2.start() 
	logging.debug("Thread t2 started.")
	t3.start()
	logging.debug("Thread t3 started.")
	t4.start()
	logging.debug("Thread t4 started.")
	t1.join(5)
	logging.debug("Thread t1 joined.")
	t2.join(5)
	logging.debug("Thread t2 joined.")
	t3.join(5)
	logging.debug("Thread t3 joined.")
	t4.join(5)
	logging.debug("Thread t4 joined.")
	logging.debug("Back to main thread.")
	
if __name__== '__main__':
	main()
	